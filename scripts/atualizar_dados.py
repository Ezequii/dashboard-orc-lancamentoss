from pathlib import Path
from datetime import datetime, date
import hashlib
import json
import shutil
import warnings
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
)

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "atualizar_dados" / "CONTROLE_DE_REQUISICOES_2026.xlsx"
OUTPUT = ROOT / "public" / "data"
BACKUPS = ROOT / "backups"
SHEET_NAME = "Acompanhamento RC 2026"

OUTPUT.mkdir(parents=True, exist_ok=True)
BACKUPS.mkdir(parents=True, exist_ok=True)

if not EXCEL.exists():
    raise SystemExit(f"ERRO: planilha nao encontrada: {EXCEL}")

stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
for filename in ("orcamentos.json", "meta.json"):
    current = OUTPUT / filename
    if current.exists():
        shutil.copy2(current, BACKUPS / f"{stamp}_{filename}")

workbook = load_workbook(EXCEL, data_only=True, read_only=True)
if SHEET_NAME not in workbook.sheetnames:
    raise SystemExit(f"ERRO: aba '{SHEET_NAME}' nao encontrada.")
worksheet = workbook[SHEET_NAME]

def clean(value):
    if value in (None, "", "*", "-"):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return " ".join(value.replace("\u00a0", " ").split()).strip() or None
    return value

def number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0

def iso_date(value):
    value = clean(value)
    if not value:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value), pattern).strftime("%Y-%m-%d")
        except (TypeError, ValueError):
            pass
    return None

def normalize_status(value):
    mapping = {
        "CONCLUIDO": "Concluído",
        "CONCLUÍDO": "Concluído",
        "FALTA NF": "Falta NF",
        "FALTA O PEDIDO": "Falta pedido",
        "FALTA PEDIDO": "Falta pedido",
        "FALTA LANCAMENTO": "Falta lançamento",
        "FALTA LANÇAMENTO": "Falta lançamento",
    }
    return mapping.get(str(value or "").strip().upper(), clean(value) or "Não informado")

records = []
for row_number, row in enumerate(
    worksheet.iter_rows(min_row=3, max_col=18, values_only=True), start=3
):
    if not clean(row[16]):
        continue
    (
        recebimento, lancamento, prefixo, equipamento, fornecedor, orcamento,
        valor_servico, valor_pecas, valor_total, solicitante, ordem_servico,
        requisicao, pedido, data_pedido, nf, data_nf, status, observacoes,
    ) = row
    records.append({
        "id": row_number - 2,
        "recebimento": iso_date(recebimento),
        "lancamento": iso_date(lancamento),
        "prefixo": clean(prefixo),
        "equipamento": clean(equipamento),
        "fornecedor": clean(fornecedor),
        "orcamento": clean(orcamento),
        "valorServico": number(valor_servico),
        "valorPecas": number(valor_pecas),
        "valorTotal": number(valor_total) if clean(valor_total) is not None else number(valor_servico) + number(valor_pecas),
        "solicitante": clean(solicitante),
        "ordemServico": clean(ordem_servico),
        "requisicao": clean(requisicao),
        "pedido": clean(pedido),
        "dataPedido": iso_date(data_pedido),
        "nf": clean(nf),
        "dataNF": iso_date(data_nf),
        "status": normalize_status(status),
        "observacoes": clean(observacoes),
    })

if not records:
    raise SystemExit("ERRO: nenhum registro valido foi processado.")

payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
checksum = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
(OUTPUT / "orcamentos.json").write_text(payload, encoding="utf-8")

now = datetime.now()
meta = {
    "atualizadoEm": now.isoformat(timespec="seconds"),
    "atualizadoEmTexto": now.strftime("%d/%m/%Y as %H:%M"),
    "arquivo": EXCEL.name,
    "linhasProcessadas": len(records),
    "checksum": checksum,
    "origem": "Atualizacao manual no computador",
}
(OUTPUT / "meta.json").write_text(
    json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)

print(f"OK_REGISTROS={len(records)}")
print(f"OK_CHECKSUM={checksum}")
print("OK_ARQUIVOS=public/data/orcamentos.json;public/data/meta.json")
